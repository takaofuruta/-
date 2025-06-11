import streamlit as st
import pandas as pd
from PyPDF2 import PdfReader
import os

# Streamlit UI
st.title("建築データ処理アプリ")
st.write("複数のPDFをアップロードし、それぞれ適切な処理を実施した後、Excelファイルを更新します。")

# PDFファイルのアップロード（複数対応）
uploaded_files = st.file_uploader("PDFをアップロード（複数選択可）", type=["pdf"], accept_multiple_files=True)

# Excelファイルのテンプレート
excel_template = "建築工事届.xlsx"

# ファイルの処理を開始
if uploaded_files:
    updated_excel = "処理済_建築工事届.xlsx"

    # Excelの元データを保持
    if os.path.exists(excel_template):
        df_excel = pd.ExcelFile(excel_template)  # 既存シート構成を維持
    else:
        st.error(f"テンプレートのExcel ({excel_template}) が見つかりません！")
        st.stop()

    # アップロードされたファイルごとに処理を実行
    for uploaded_file in uploaded_files:
        pdf_name = uploaded_file.name  # ファイル名を取得
        st.write(f"📂 アップロードされたファイル: {pdf_name}")

        try:
            reader = PdfReader(uploaded_file)
            extracted_text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])

            # ファイル名に応じて処理を分岐
            if "図面データ.pdf" in pdf_name:
                st.write(f"⚙ {pdf_name} → **図面データ処理開始**")
                # ここで図面データの処理を追加

            elif "面積表　図面.pdf" in pdf_name:
                st.write(f"⚙ {pdf_name} → **面積表データ処理開始**")
                # ここで面積情報抽出処理を追加

            else:
                st.warning(f"⚠ {pdf_name} は対応する処理がありません。スキップします。")

        except Exception as e:
            st.error(f"❌ PDF処理中にエラーが発生: {e}")

    # Excelの更新（元の構造を維持）
    with pd.ExcelWriter(updated_excel, mode="w", engine="openpyxl") as writer:
        for sheet_name in df_excel.sheet_names:
            df = pd.read_excel(df_excel, sheet_name=sheet_name)  # 修正: `df_excel` ではなく `pd.read_excel` を使用
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    st.success(f"✅ Excelファイル ({updated_excel}) を更新しました！")

    # ダウンロードボタン
    with open(updated_excel, "rb") as f:
        st.download_button(label="📥 処理済みExcelをダウンロード", data=f, file_name=updated_excel, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






from PyPDF2 import PdfReader
from pdfminer.high_level import extract_text

# PDFを直接読み込む
reader = PdfReader("図面データ.pdf")  
full_text = reader.pages[0].extract_text()

# PDFからテキストを抽出
text = extract_text("図面データ.pdf")

# 各種情報を抽出するための処理
name_start = text.find("建築主:") + len("建築主:")
name_end = text.find("〒") - 1
name = text[name_start:name_end]

yubinbango_start = text.find("〒") + 1 
yubinbango_end = text.find("-", yubinbango_start) 
yubinbango = text[yubinbango_start:yubinbango_end]

hyphen_index = text.find("-")
yubinbango_last4 = text[hyphen_index + 1: hyphen_index + 5]

zip_pos = text.find("〒")
next_line_start = text.find("\n", zip_pos) + 1
address_end = text.find("建築場所（地名地番）")
top_address = text[next_line_start:address_end].strip()

place_start = text.find("建築場所（地名地番）")
place_line_start = text.find("\n", place_start) + 1
place_line_end = text.find("\n", place_line_start)
place_line = text[place_line_start:place_line_end].strip()

ken_index = place_line.find("県")
ken_name = place_line[:ken_index + 1] if ken_index != -1 else ""

start_pos = text.find("建築場所（地名地番）")
end_pos = text.find("建築場所（住居表示）")
start_pos += len("建築場所（地名地番）")

address = text[start_pos:end_pos].strip()

lines = text.splitlines()

kouji_line = ""
for line in lines:
    line_stripped = line.strip()
    if line_stripped.endswith("工 事"):
        kouji_line = line_stripped
        break

# 抽出した情報の出力
print("建築主:", name)
print("郵便番号:", yubinbango)
print("郵便番号（後ろ4桁）:", yubinbango_last4)
print("住所:", top_address)
print("建築場所:", address)
print("工事名:", kouji_line)

import re

lines = text.splitlines()

# 面積項目リスト（順番通り）
area_labels = [
    "敷地面積",       # → site_area
    "1階床面積",
    "2階床面積",
    "延床面積",       # → total_floor_area
    "建築面積",
    "建蔽率",
    "容積率"
]

numeric_values = []
for line in lines:
    cleaned = line.strip().replace("㎡", "").replace("％", "")
    if re.fullmatch(r"[\d.]+", cleaned):
        numeric_values.append(cleaned)

site_area = numeric_values[0]             
total_floor_area = numeric_values[3]      

import openpyxl
from openpyxl.styles import Alignment

excel_path = "建築工事届.xlsx"
sheet_name = "建築工事届（別記第40号様式）"
sheet_name2 = "第三種換気"

wb = openpyxl.load_workbook(excel_path)

# シート1: "建築工事届（別記第40号様式）"
ws = wb[sheet_name]

ken_cells = ["A12", "J72"]
for ken_cell in ken_cells:
    ws[ken_cell] = ken_name
ws["I16"] = name
ws["I17"] = yubinbango
ws["O17"] = yubinbango_last4
ws["I18"] = top_address
ws["O72"] = address
ws["J85"] = kouji_line
ws["K92"] = total_floor_area
ws["S109"] = site_area

# シート2: "第三種換気"
ws2 = wb[sheet_name2]  # シートを取得

# B1セルに値を入力
ws2["B1"] = kouji_line
# B1セルを右揃えに設定 (こちらでシート"ws2"を正しく指定)
ws2["B1"].alignment = Alignment(horizontal="right", vertical="center")

# 保存処理
wb.save(excel_path)




import re
from PyPDF2 import PdfReader
import openpyxl
import unicodedata

# PDFからテキストを抽出
reader = PdfReader("面積表　図面.pdf")
full_text = reader.pages[0].extract_text()

# 前処理：空白や改行を削除
processed_text = re.sub(r"\s+", "", full_text)

# 表を識別するパターン（「部屋名」から「合計」までを検出）
table_pattern = r"部屋名.*?合計"
tables = re.findall(table_pattern, processed_text)

# 部屋名と面積を抽出する正規表現パターン
room_pattern = r"(玄関|階段|トイレ|ホール|ＬＤＫ|洗面脱衣室|洋室|廊下|サービスルーム)[^\d]*([\d.]+)"

# 平均天井高の値
ceiling_heights = {
    "玄関": 2.58,
    "ホール": 2.4,
    "階段": 2.875,
    "洗面脱衣室": 2.4,
    "洋室": 2.4,
    "トイレ": 2.4,
    "廊下": 2.4,
    "サービスルーム": 2.4,
    "ＬＤＫ": 2.4
}

# 各階のデータを格納
floor_data = {}
for table in tables:
    # 「玄関」があるかどうかで1階と2階を判断
    if "玄関" in table:
        floor_name = "1階"
    else:
        floor_name = "2階"

    matches = re.findall(room_pattern, table)
    
    # 部屋名のユニーク化（番号付けを追加）
    room_counts = {}
    floor_areas = {}
    for room, area in matches:
        room_counts[room] = room_counts.get(room, 0) + 1
        unique_room = f"{room}{room_counts[room]}" if room_counts[room] > 1 else room
        floor_areas[unique_room] = float(area)
    
    # 各階のデータに追加
    if floor_name not in floor_data:
        floor_data[floor_name] = floor_areas
    else:
        # 同じ階にデータが追加される場合
        floor_data[floor_name].update(floor_areas)

# エクセルファイルを開く
excel_path = "建築工事届.xlsx"
sheet_name2 = "第三種換気"
workbook = openpyxl.load_workbook(excel_path)

# 指定したシートを取得
if sheet_name2 in workbook.sheetnames:
    sheet = workbook[sheet_name2]
else:
    raise ValueError(f"指定されたシート名 '{sheet_name2}' が見つかりません！")

# 転記処理を実施
def normalize_text(text):
    if text:
        return unicodedata.normalize("NFKC", text).strip().upper()
    return ""

# 1階部分のデータを転記 (D8:D18, B列に室名, E列に平均天井高)
ground_floor_areas = floor_data.get("1階", {})
for i, (room, area) in enumerate(ground_floor_areas.items(), start=8):
    sheet.cell(row=i, column=2, value=room)  # 室名をB列に記入
    sheet.cell(row=i, column=4, value=area)  # 面積をD列に記入
    # 平均天井高をE列に記入（室名から値を取得）
    room_base_name = re.sub(r"\d+$", "", room)  # ユニーク化された番号を除去
    sheet.cell(row=i, column=5, value=ceiling_heights.get(room_base_name, "-"))  # 対応がない場合は "-"

# 2階部分のデータを転記 (D19:D28, B列に室名, E列に平均天井高)
first_floor_areas = floor_data.get("2階", {})
for i, (room, area) in enumerate(first_floor_areas.items(), start=19):
    sheet.cell(row=i, column=2, value=room)  # 室名をB列に記入
    sheet.cell(row=i, column=4, value=area)  # 面積をD列に記入
    # 平均天井高をE列に記入（室名から値を取得）
    room_base_name = re.sub(r"\d+$", "", room)  # ユニーク化された番号を除去
    sheet.cell(row=i, column=5, value=ceiling_heights.get(room_base_name, "-"))  # 対応がない場合は "-"

# エクセルファイルを上書き保存
workbook.save(excel_path)

print("1階・2階のデータ転記が完了しました！（B列に室名、E列に平均天井高対応済み）")




# エクセルファイルを開く
excel_path = "建築工事届.xlsx"
sheet_name2 = "第三種換気"
workbook = openpyxl.load_workbook(excel_path)

# 指定したシートを取得
if sheet_name2 in workbook.sheetnames:
    sheet = workbook[sheet_name2]
else:
    raise ValueError(f"指定されたシート名 '{sheet_name2}' が見つかりません！")

# 転記処理を実施
def normalize_text(text):
    if text:
        return unicodedata.normalize("NFKC", text).strip().upper()
    return ""

# 1階部分のデータを転記 (D8:D18, B列に室名)
ground_floor_areas = floor_data.get("1階", {})
for i, (room, area) in enumerate(ground_floor_areas.items(), start=8):
    sheet.cell(row=i, column=2, value=room)  # 室名をB列に記入
    sheet.cell(row=i, column=4, value=area)  # 面積をD列に記入

# 2階部分のデータを転記 (D19:D28, B列に室名)
first_floor_areas = floor_data.get("2階", {})
for i, (room, area) in enumerate(first_floor_areas.items(), start=19):
    sheet.cell(row=i, column=2, value=room)  # 室名をB列に記入
    sheet.cell(row=i, column=4, value=area)  # 面積をD列に記入

# エクセルファイルを上書き保存
workbook.save(excel_path)

print("1階・2階のデータ転記が完了しました！（順序修正済み）")
